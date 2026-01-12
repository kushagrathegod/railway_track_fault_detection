from fastapi import FastAPI, Depends, HTTPException, BackgroundTasks, status, UploadFile, File
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from fastapi.responses import StreamingResponse
from fastapi.staticfiles import StaticFiles
from dotenv import load_dotenv

load_dotenv()
from fastapi.middleware.cors import CORSMiddleware
from sqlalchemy.orm import text, Session
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta
import subprocess
import os
from pathlib import Path
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import pytz
import database
from database import Defect, SessionLocal, User, Station
import groq_service
import email_service
import auth
from location_utils import find_nearest_station

app = FastAPI(title="Railway Defect Detection System")

# OAuth2 scheme for JWT
oauth2_scheme = OAuth2PasswordBearer(tokenUrl="auth/login")

# CORS Setup
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Mount static files for images
uploads_dir = Path(__file__).parent / "uploads"
uploads_dir.mkdir(exist_ok=True)
app.mount("/uploads", StaticFiles(directory=str(uploads_dir)), name="uploads")

# Database Dependency
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

# Initialize API
@app.on_event("startup")
def startup():
    database.init_db()
    db = SessionLocal()
    try:
        if database.DATABASE_URL.startswith("postgresql"):
            resync_sql = """
            SELECT setval(pg_get_serial_sequence('stations', 'id'), coalesce(max(id), 1), max(id) IS NOT NULL) FROM stations;
            SELECT setval(pg_get_serial_sequence('users', 'id'), coalesce(max(id), 1), max(id) IS NOT NULL) FROM users;
            SELECT setval(pg_get_serial_sequence('defects', 'id'), coalesce(max(id), 1), max(id) IS NOT NULL) FROM defects;
            """
            for statement in resync_sql.strip().split(';'):
                if statement.strip():
                    db.execute(text(statement))
            db.commit()
            print("✅ ID sequences resynced for PostgreSQL")
        
        admin_exists = db.query(User).filter(User.role == "Admin").first()
        if not admin_exists:
            print("🚀 No Admin found in PostgreSQL. Seeding default Admin...")
            from auth import get_password_hash
            admin = User(
                username="admin",
                email="kush85114@gmail.com",
                hashed_password=get_password_hash("admin123"), # You can change this
                role="Admin"
            )
            db.add(admin)
            db.commit()
            print("✅ Default Admin created: admin / admin123")
            
    except Exception as e:
        print(f"❌ Seeding error: {e}")
    finally:
        db.close()

# Auth Dependencies
def get_current_user(token: str = Depends(oauth2_scheme), db: Session = Depends(get_db)):
    """Get the current authenticated user from JWT token."""
    credentials_exception = HTTPException(
        status_code=status.HTTP_401_UNAUTHORIZED,
        detail="Could not validate credentials",
        headers={"WWW-Authenticate": "Bearer"},
    )
    
    payload = auth.decode_access_token(token)
    if payload is None:
        raise credentials_exception
    
    username: str = payload.get("sub")
    if username is None:
        raise credentials_exception
    
    user = db.query(User).filter(User.username == username).first()
    if user is None:
        raise credentials_exception
    
    return user

def require_admin(current_user: User = Depends(get_current_user)):
    """Require that the current user is an admin."""
    if current_user.role != "Admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

# Pydantic Models
class UserLogin(BaseModel):
    username: str
    password: str

class Token(BaseModel):
    access_token: str
    token_type: str

class UserResponse(BaseModel):
    id: int
    username: str
    email: str
    role: str
    station_id: Optional[int]
    
    class Config:
        from_attributes = True

class StationResponse(BaseModel):
    id: int
    name: str
    code: str
    latitude: float
    longitude: float
    station_master_email: str
    
    class Config:
        from_attributes = True

class StationCreate(BaseModel):
    name: str
    code: str
    latitude: float
    longitude: float
    station_master_email: str
    station_master_username: str
    station_master_password: str

class StationUpdate(BaseModel):
    name: Optional[str] = None
    code: Optional[str] = None
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    station_master_email: Optional[str] = None

# Pydantic Models
class DefectCreate(BaseModel):
    defect_type: str
    confidence: float
    image_url: str
    latitude: Optional[float] = None
    longitude: Optional[float] = None
    nearest_station: Optional[str] = None

class DefectResponse(DefectCreate):
    id: int
    severity: str
    root_cause: Optional[str]
    action_required: Optional[str]
    resolution_steps: Optional[str]
    timestamp: datetime
    assigned_station_id: Optional[int]
    status: str
    resolved_at: Optional[datetime]
    resolved_by: Optional[int]

    class Config:
        from_attributes = True

def normalize_severity(severity_str):
    """
    Normalize severity to one of: Low, High, Critical
    """
    if not severity_str:
        return "High"  # Default to High for safety
    
    severity_lower = str(severity_str).lower().strip()
    
    if severity_lower in ["critical", "severe", "very high", "urgent"]:
        return "Critical"
    elif severity_lower in ["high", "significant", "medium-high"]:
        return "High"
    elif severity_lower in ["low", "minor", "minimal"]:
        return "Low"
    elif severity_lower in ["medium", "moderate"]:
        return "High"  # Treat medium as High for safety
    else:
        return "High"  # Default to High for unknown values

# Web Upload Endpoint for Drone Control Page (Uses Vision Agent ML Pipeline)
@app.post("/upload-analyze")
async def upload_and_analyze(
    file: UploadFile = File(...),
    latitude: Optional[float] = None,
    longitude: Optional[float] = None,
    background_tasks: BackgroundTasks = None,
    current_user: User = Depends(get_current_user),
    db: Session = Depends(get_db)
):
    """
    Web upload endpoint that uses the same pipeline as vision_agent.py:
    1. Call ML Model API (EfficientNet) for defect detection
    2. Check confidence threshold (70%)
    3. If defective and confident → Groq analysis and save to DB
    
    Accepts optional latitude/longitude parameters for manual location input.
    Defaults to New Delhi (28.6139, 77.2090) if not provided.
    """
    try:
        import requests
        import io
        from PIL import Image
        
        # Read file
        contents = await file.read()
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"web_upload_{timestamp}.jpg"
        upload_dir = "uploads"
        os.makedirs(upload_dir, exist_ok=True)
        filepath = os.path.join(upload_dir, filename)
        
        # Save file
        with open(filepath, "wb") as f:
            f.write(contents)
        
        # STEP 1: Call ML Model API (same as vision agent)
        MODEL_API_URL = os.getenv("MODEL_API_URL", "https://vishalbhagat01-railway.hf.space/predict")
        CONFIDENCE_THRESHOLD = 70.0
        
        
        # Prepare file for ML model
        files = {'file': ('image.jpg', contents, 'image/jpeg')}
        
        try:
            ml_response = requests.post(MODEL_API_URL, files=files, timeout=30)
            ml_data = ml_response.json()
            
            prediction = ml_data.get("prediction", "Unknown")
            confidence = float(ml_data.get("confidence", 0))
            
            
            # STEP 2: Check if defective and above threshold
            if prediction != "Defective":
                return {
                    "status": "no_defect",
                    "prediction": prediction,
                    "confidence": confidence,
                    "message": "No defect detected by ML model"
                }
            
            if confidence < CONFIDENCE_THRESHOLD:
                return {
                    "status": "low_confidence",
                    "prediction": prediction,
                    "confidence": confidence,
                    "message": f"Confidence too low ({confidence}%). Threshold is {CONFIDENCE_THRESHOLD}%"
                }
            
            # STEP 3: Defect detected! Proceed with Groq analysis
            
            # Use provided location or default to New Delhi
            location_data = {
                "latitude": latitude if latitude is not None else 28.6139,
                "longitude": longitude if longitude is not None else 77.2090,
                "nearest_station": "Web Upload"
            }
            
            location_str = f"Lat: {location_data['latitude']}, Lon: {location_data['longitude']}, Station: {location_data['nearest_station']}"
            
            # Groq Analysis
            import asyncio
            try:
                analysis = await asyncio.to_thread(
                    groq_service.analyze_defect,
                    "Track Defect",
                    confidence,
                    location_str
                )
            except Exception as e:
                analysis = {}
            
            severity = normalize_severity(analysis.get("severity", "High"))
            
            # Create defect record
            # Store relative path for serving via static files
            filename = os.path.basename(filepath)
            db_defect = Defect(
                defect_type="Track Defect",
                confidence=confidence,
                image_url=f"/uploads/{filename}",
                latitude=location_data["latitude"],
                longitude=location_data["longitude"],
                nearest_station=location_data["nearest_station"],
                severity=severity,
                root_cause=str(analysis.get("root_cause", "Analysis pending")),
                action_required=str(analysis.get("immediate_action", "Awaiting assessment")),
                resolution_steps=str(analysis.get("resolution_steps", "Pending analysis"))
            )
            
            db.add(db_defect)
            db.commit()
            db.refresh(db_defect)
            
            # Assign to nearest station
            nearest_station = find_nearest_station(
                location_data["latitude"],
                location_data["longitude"],
                db
            )
            if nearest_station:
                db_defect.assigned_station_id = nearest_station.id
                db.commit()
                db.refresh(db_defect)
                
                # Send email if critical
                if severity == "Critical" and background_tasks:
                    defect_dict = {
                        "defect_type": db_defect.defect_type,
                        "confidence": db_defect.confidence,
                        "latitude": db_defect.latitude,
                        "longitude": db_defect.longitude,
                        "nearest_station": db_defect.nearest_station,
                        "timestamp": db_defect.timestamp,
                        "severity": db_defect.severity,
                        "action_required": db_defect.action_required,
                        "resolution_steps": db_defect.resolution_steps,
                        "image_url": db_defect.image_url
                    }
                    background_tasks.add_task(
                        email_service.send_alert,
                        defect_dict,
                        recipient_email=nearest_station.station_master_email,
                        station_name=nearest_station.name
                    )
            
            # Return defect details
            return {
                "status": "defect_detected",
                "prediction": prediction,
                "confidence": confidence,
                "id": db_defect.id,
                "defect_type": db_defect.defect_type,
                "severity": db_defect.severity,
                "root_cause": db_defect.root_cause,
                "action_required": db_defect.action_required,
                "resolution_steps": db_defect.resolution_steps,
                "timestamp": db_defect.timestamp.isoformat(),
                "assigned_station_id": db_defect.assigned_station_id
            }
            
        except requests.exceptions.Timeout:
            raise HTTPException(status_code=504, detail="ML Model API timeout")
        except requests.exceptions.ConnectionError:
            raise HTTPException(status_code=503, detail="Cannot connect to ML Model API")
        except Exception as e:
            print(f"ML Model API error: {e}")
            raise HTTPException(status_code=500, detail=f"ML Model error: {str(e)}")
        
    except Exception as e:
        print(f"Upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@app.post("/analyze", response_model=DefectResponse)
async def analyze_defect(defect: DefectCreate, background_tasks: BackgroundTasks, db: Session = Depends(get_db)):
    """
    Receives detection data from Vision System.
    1. Calls Groq for analysis.
    2. Saves to DB.
    3. Sends alert if Critical.
    """
    location_str = f"Lat: {defect.latitude}, Lon: {defect.longitude}, Station: {defect.nearest_station}"
    
    # 1. Groq Analysis
    print(f"Analyzing defect: {defect.defect_type} with confidence {defect.confidence}%")
    # Offload blocking synchronous call to thread pool
    import asyncio
    try:
        analysis = await asyncio.to_thread(groq_service.analyze_defect, defect.defect_type, defect.confidence, location_str)
    except Exception as e:
        print(f"ERROR calling Groq Service: {e}")
        analysis = {}

    # Normalize severity
    severity = normalize_severity(analysis.get("severity", "High"))
    
    # 2. Save to DB
    db_defect = Defect(
        defect_type=defect.defect_type,
        confidence=defect.confidence,
        image_url=defect.image_url,
        latitude=defect.latitude,
        longitude=defect.longitude,
        nearest_station=defect.nearest_station,
        severity=severity,
        root_cause=str(analysis.get("root_cause", "Analysis pending")) if not isinstance(analysis.get("root_cause"), list) else "; ".join(analysis.get("root_cause")),
        action_required=str(analysis.get("immediate_action", "Awaiting assessment")) if not isinstance(analysis.get("immediate_action"), list) else "; ".join(analysis.get("immediate_action")),
        resolution_steps=str(analysis.get("resolution_steps", "Pending detailed analysis")) if not isinstance(analysis.get("resolution_steps"), list) else "; ".join(analysis.get("resolution_steps"))
    )
    db.add(db_defect)
    db.commit()
    db.refresh(db_defect)
    
    print(f"Defect saved to DB with ID: {db_defect.id}, Severity: {db_defect.severity}")
    
    # 3. Find nearest station and assign
    nearest_station = None
    if defect.latitude and defect.longitude:
        nearest_station = find_nearest_station(defect.latitude, defect.longitude, db)
        if nearest_station:
            db_defect.assigned_station_id = nearest_station.id
            db.commit()
            db.refresh(db_defect)
            print(f"Defect assigned to station: {nearest_station.name} ({nearest_station.code})")
    
    # 4. Background Task for Email to Station Master
    if db_defect.severity == "Critical":
        print(f"Critical defect detected! Sending email alert...")
        defect_dict = {
            "defect_type": db_defect.defect_type,
            "confidence": db_defect.confidence,
            "latitude": db_defect.latitude,
            "longitude": db_defect.longitude,
            "nearest_station": db_defect.nearest_station,
            "timestamp": db_defect.timestamp,
            "severity": db_defect.severity,
            "action_required": db_defect.action_required,
            "resolution_steps": db_defect.resolution_steps,
            "image_url": db_defect.image_url
        }
        
        # Send to station master if station assigned
        if nearest_station:
            background_tasks.add_task(
                email_service.send_alert, 
                defect_dict, 
                recipient_email=nearest_station.station_master_email,
                station_name=nearest_station.name
            )
        else:
            # Fallback to default recipient if no station found
            background_tasks.add_task(email_service.send_alert, defect_dict)
        
    return db_defect

@app.get("/defects", response_model=List[DefectResponse])
def get_defects(skip: int = 0, limit: int = 100, db: Session = Depends(get_db)):
    defects = db.query(Defect).order_by(Defect.timestamp.desc()).offset(skip).limit(limit).all()
    return defects

@app.get("/defects/export/excel")
def export_defects_excel(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """
    Export all defects to Excel file.
    Available for both Admin and StationMaster users.
    """
    # Get all defects
    defects = db.query(Defect).order_by(Defect.timestamp.desc()).all()
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Defect Reports"
    
    # Define IST timezone
    ist = pytz.timezone('Asia/Kolkata')
    
    # Define headers
    headers = [
        "ID", "Timestamp (IST)", "Defect Type", "Severity", "Status", 
        "Confidence (%)", "Latitude", "Longitude", "Nearest Station",
        "Root Cause", "Action Required", "Resolution Steps", 
        "Resolved At (IST)", "Resolved By"
    ]
    
    # Add headers with styling
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1F4788", end_color="1F4788", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Add data rows
    for row_num, defect in enumerate(defects, 2):
        # Convert timestamp to IST
        timestamp_ist = defect.timestamp.replace(tzinfo=pytz.UTC).astimezone(ist) if defect.timestamp else None
        resolved_at_ist = defect.resolved_at.replace(tzinfo=pytz.UTC).astimezone(ist) if defect.resolved_at else None
        
        # Get resolver username
        resolver_name = ""
        if defect.resolved_by:
            resolver = db.query(User).filter(User.id == defect.resolved_by).first()
            resolver_name = resolver.username if resolver else f"User ID {defect.resolved_by}"
        
        row_data = [
            defect.id,
            timestamp_ist.strftime("%Y-%m-%d %H:%M:%S") if timestamp_ist else "",
            defect.defect_type or "",
            defect.severity or "",
            defect.status or "Open",
            defect.confidence or 0,
            defect.latitude or "",
            defect.longitude or "",
            defect.nearest_station or "",
            defect.root_cause or "",
            defect.action_required or "",
            defect.resolution_steps or "",
            resolved_at_ist.strftime("%Y-%m-%d %H:%M:%S") if resolved_at_ist else "",
            resolver_name
        ]
        
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            
            # Color code severity column
            if col_num == 4:  # Severity column
                if value == "Critical":
                    cell.fill = PatternFill(start_color="FFD7D7", end_color="FFD7D7", fill_type="solid")
                    cell.font = Font(color="C00000", bold=True)
                elif value == "High":
                    cell.fill = PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid")
                    cell.font = Font(color="E67300", bold=True)
                elif value == "Low":
                    cell.fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                    cell.font = Font(color="997300", bold=True)
    
    # Auto-size columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)  # Cap at 50 for very long text
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to BytesIO
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    
    # Generate filename with current date
    filename = f"railway_defects_report_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    
    # Return as streaming response
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.patch("/defects/{defect_id}/resolve", response_model=DefectResponse)
def resolve_defect(
    defect_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Mark a defect as resolved. Station masters can only resolve defects assigned to their station."""
    defect = db.query(Defect).filter(Defect.id == defect_id).first()
    
    if not defect:
        raise HTTPException(status_code=404, detail="Defect not found")
    
    # Check authorization
    if current_user.role == "StationMaster":
        # Station masters can only resolve defects assigned to their station
        if defect.assigned_station_id != current_user.station_id:
            raise HTTPException(
                status_code=403, 
                detail="You can only resolve defects assigned to your station"
            )
    # Admins can resolve any defect
    
    # Check if already resolved
    if defect.status == "Resolved":
        raise HTTPException(status_code=400, detail="Defect is already marked as resolved")
    
    # Mark as resolved
    defect.status = "Resolved"
    defect.resolved_at = datetime.now()
    defect.resolved_by = current_user.id
    
    db.commit()
    db.refresh(defect)
    
    return defect

@app.patch("/defects/{defect_id}/reopen", response_model=DefectResponse)
def reopen_defect(
    defect_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)
):
    """Reopen a resolved defect (Admin only)."""
    defect = db.query(Defect).filter(Defect.id == defect_id).first()
    
    if not defect:
        raise HTTPException(status_code=404, detail="Defect not found")
    
    # Check if already open
    if defect.status == "Open":
        raise HTTPException(status_code=400, detail="Defect is already open")
    
    # Reopen the defect
    defect.status = "Open"
    defect.resolved_at = None
    defect.resolved_by = None
    
    db.commit()
    db.refresh(defect)
    
    return defect

@app.delete("/defects/{defect_id}")
def delete_defect(
    defect_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)
):
    """Delete a defect record permanently (Admin only)."""
    defect = db.query(Defect).filter(Defect.id == defect_id).first()
    
    if not defect:
        raise HTTPException(status_code=404, detail="Defect not found")
    
    # Optionally delete the image file if it exists
    if defect.image_url and defect.image_url.startswith("/uploads/"):
        try:
            filename = defect.image_url.replace("/uploads/", "")
            filepath = os.path.join("uploads", filename)
            if os.path.exists(filepath):
                os.remove(filepath)
        except Exception as e:
            print(f"Error deleting image file: {e}")
            # Continue with database deletion even if file deletion fails
    
    # Delete from database
    db.delete(defect)
    db.commit()
    
    return {"message": "Defect deleted successfully", "defect_id": defect_id}

class BulkDeleteRequest(BaseModel):
    defect_ids: List[int]

@app.post("/defects/bulk-delete")
def bulk_delete_defects(
    request: BulkDeleteRequest,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)
):
    """Delete multiple defect records permanently (Admin only)."""
    if not request.defect_ids:
        raise HTTPException(status_code=400, detail="No defect IDs provided")
    
    deleted_count = 0
    errors = []
    
    for defect_id in request.defect_ids:
        try:
            defect = db.query(Defect).filter(Defect.id == defect_id).first()
            
            if not defect:
                errors.append(f"Defect {defect_id} not found")
                continue
            
            # Optionally delete the image file if it exists
            if defect.image_url and defect.image_url.startswith("/uploads/"):
                try:
                    filename = defect.image_url.replace("/uploads/", "")
                    filepath = os.path.join("uploads", filename)
                    if os.path.exists(filepath):
                        os.remove(filepath)
                except Exception as e:
                    print(f"Error deleting image file for defect {defect_id}: {e}")
            
            # Delete from database
            db.delete(defect)
            deleted_count += 1
        except Exception as e:
            errors.append(f"Error deleting defect {defect_id}: {str(e)}")
    
    db.commit()
    
    return {
        "message": f"Successfully deleted {deleted_count} defect(s)",
        "deleted_count": deleted_count,
        "errors": errors if errors else None
    }


# Drone Inspection Control (Admin only)
drone_process = None

@app.post("/drone/start")
def start_drone_inspection(admin: User = Depends(require_admin)):
    """Start the drone inspection vision agent (Admin only)."""
    global drone_process
    
    if drone_process and drone_process.poll() is None:
        raise HTTPException(status_code=400, detail="Drone inspection is already running")
    
    try:
        # Get path to vision agent
        vision_agent_path = Path(__file__).parent.parent / "vision" / "vision_agent.py"
        
        if not vision_agent_path.exists():
            raise HTTPException(status_code=500, detail="Vision agent script not found")
        
        # Start the vision agent in non-interactive drone mode
        drone_process = subprocess.Popen(
            ["python", str(vision_agent_path), "--mode", "drone"],
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            bufsize=1,
            creationflags=subprocess.CREATE_NO_WINDOW if os.name == 'nt' else 0
        )
        
        return {
            "status": "started",
            "message": "Drone inspection started successfully",
            "process_id": drone_process.pid
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to start drone inspection: {str(e)}")

@app.post("/drone/stop")
def stop_drone_inspection(admin: User = Depends(require_admin)):
    """Stop the drone inspection vision agent (Admin only)."""
    global drone_process
    
    if not drone_process or drone_process.poll() is not None:
        raise HTTPException(status_code=400, detail="Drone inspection is not running")
    
    try:
        drone_process.terminate()
        drone_process.wait(timeout=5)
        drone_process = None
        
        return {
            "status": "stopped",
            "message": "Drone inspection stopped successfully"
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to stop drone inspection: {str(e)}")

@app.get("/drone/status")
def get_drone_status(current_user: User = Depends(get_current_user)):
    """Get the current status of drone inspection."""
    global drone_process
    
    is_running = drone_process is not None and drone_process.poll() is None
    
    return {
        "is_running": is_running,
        "process_id": drone_process.pid if is_running else None
    }

# Authentication Endpoints
@app.post("/auth/login", response_model=Token)
def login(form_data: OAuth2PasswordRequestForm = Depends(), db: Session = Depends(get_db)):
    """Login endpoint - returns JWT token."""
    user = db.query(User).filter(User.username == form_data.username).first()
    
    if not user or not auth.verify_password(form_data.password, user.hashed_password):
        raise HTTPException(
            status_code=status.HTTP_401_UNAUTHORIZED,
            detail="Incorrect username or password",
            headers={"WWW-Authenticate": "Bearer"},
        )
    
    access_token_expires = timedelta(minutes=auth.ACCESS_TOKEN_EXPIRE_MINUTES)
    access_token = auth.create_access_token(
        data={"sub": user.username}, expires_delta=access_token_expires
    )
    
    return {"access_token": access_token, "token_type": "bearer"}

@app.get("/auth/me", response_model=UserResponse)
def get_me(current_user: User = Depends(get_current_user)):
    """Get current user information."""
    return current_user

# Station Endpoints
@app.get("/stations", response_model=List[StationResponse])
def get_stations(db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    """Get all railway stations."""
    stations = db.query(Station).all()
    return stations

@app.post("/stations", response_model=StationResponse, status_code=status.HTTP_201_CREATED)
def create_station(
    station: StationCreate, 
    db: Session = Depends(get_db), 
    admin: User = Depends(require_admin)
):
    """Create a new railway station (Admin only)."""
    # Check if station with same name or code already exists
    existing_name = db.query(Station).filter(Station.name == station.name).first()
    if existing_name:
        raise HTTPException(status_code=400, detail=f"Station with name '{station.name}' already exists")
    
    existing_code = db.query(Station).filter(Station.code == station.code.upper()).first()
    if existing_code:
        raise HTTPException(status_code=400, detail=f"Station with code '{station.code}' already exists")
    
    # Validate coordinates
    if not (-90 <= station.latitude <= 90):
        raise HTTPException(status_code=400, detail="Latitude must be between -90 and 90")
    if not (-180 <= station.longitude <= 180):
        raise HTTPException(status_code=400, detail="Longitude must be between -180 and 180")
    
    # Check if user with same username or email already exists
    existing_user = db.query(User).filter(
        (User.username == station.station_master_username) | 
        (User.email == station.station_master_email)
    ).first()
    if existing_user:
        raise HTTPException(
            status_code=400, 
            detail=f"User with username '{station.station_master_username}' or email '{station.station_master_email}' already exists"
        )

    # Create new station
    db_station = Station(
        name=station.name,
        code=station.code.upper(),
        latitude=station.latitude,
        longitude=station.longitude,
        station_master_email=station.station_master_email
    )
    db.add(db_station)
    db.commit()
    db.refresh(db_station)
    
    # Create corresponding Station Master user
    db_user = User(
        username=station.station_master_username,
        email=station.station_master_email,
        hashed_password=auth.get_password_hash(station.station_master_password),
        role="StationMaster",
        station_id=db_station.id
    )
    db.add(db_user)
    db.commit()
    
    return db_station

@app.put("/stations/{station_id}", response_model=StationResponse)
def update_station(
    station_id: int,
    station_update: StationUpdate,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)
):
    """Update a railway station (Admin only)."""
    db_station = db.query(Station).filter(Station.id == station_id).first()
    if not db_station:
        raise HTTPException(status_code=404, detail="Station not found")
    
    # Check for conflicts if name or code is being updated
    if station_update.name is not None and station_update.name != db_station.name:
        existing = db.query(Station).filter(
            Station.name == station_update.name,
            Station.id != station_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail=f"Station with name '{station_update.name}' already exists")
        db_station.name = station_update.name
    
    if station_update.code is not None and station_update.code.upper() != db_station.code:
        existing = db.query(Station).filter(
            Station.code == station_update.code.upper(),
            Station.id != station_id
        ).first()
        if existing:
            raise HTTPException(status_code=400, detail=f"Station with code '{station_update.code}' already exists")
        db_station.code = station_update.code.upper()
    
    # Update coordinates if provided
    if station_update.latitude is not None:
        if not (-90 <= station_update.latitude <= 90):
            raise HTTPException(status_code=400, detail="Latitude must be between -90 and 90")
        db_station.latitude = station_update.latitude
    
    if station_update.longitude is not None:
        if not (-180 <= station_update.longitude <= 180):
            raise HTTPException(status_code=400, detail="Longitude must be between -180 and 180")
        db_station.longitude = station_update.longitude
    
    # Update email if provided
    if station_update.station_master_email is not None:
        db_station.station_master_email = station_update.station_master_email
    
    db.commit()
    db.refresh(db_station)
    
    return db_station

@app.delete("/stations/{station_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_station(
    station_id: int,
    db: Session = Depends(get_db),
    admin: User = Depends(require_admin)
):
    """Delete a railway station (Admin only)."""
    print(f"DELETION_REQUEST: Station {station_id} requested by {admin.username}")
    
    db_station = db.query(Station).filter(Station.id == station_id).first()
    if not db_station:
        print(f"DELETION_FAILED: Station {station_id} not found")
        raise HTTPException(status_code=404, detail="Station not found")
    
    # Check if station has assigned defects
    assigned_defects = db.query(Defect).filter(Defect.assigned_station_id == station_id).count()
    if assigned_defects > 0:
        print(f"DELETION_FAILED: Station {station_id} has {assigned_defects} assigned defects")
        raise HTTPException(
            status_code=400, 
            detail=f"Cannot delete station with {assigned_defects} assigned defects. Please reassign or delete them first."
        )
    
    try:
        # Delete associated users and the station in a single transaction
        print(f"DELETION_EXECUTE: Removing associated users for station {station_id}")
        db.query(User).filter(User.station_id == station_id).delete()
        
        print(f"DELETION_EXECUTE: Removing station record {station_id}")
        db.delete(db_station)
        
        db.commit()
        print(f"DELETION_SUCCESS: Station {station_id} and its associated users removed")
    except Exception as e:
        db.rollback()
        print(f"DELETION_ERROR: {str(e)}")
        raise HTTPException(status_code=500, detail=f"Database error during deletion: {str(e)}")
    
    return None

@app.get("/")
def read_root():
    return {"status": "ok", "message": "Railway Defect Detection API is running"}
